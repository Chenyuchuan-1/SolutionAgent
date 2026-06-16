from __future__ import annotations

import csv
import json
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from app.core.logging import get_logger
from app.services.storage import RunStorage
from tools.document.mineru_parser import MinerUParser

logger = get_logger(__name__)


class DocumentService:
    def __init__(self) -> None:
        self.storage = RunStorage()
        self.mineru = MinerUParser()

    async def parse_run_uploads(self, run_id: str, uploaded_files: list[str] | None = None) -> list[dict[str, Any]]:
        meta = self.storage.load_run(run_id)
        candidates = meta.get("uploaded_files", [])
        if uploaded_files:
            names = set(uploaded_files)
            candidates = [f for f in candidates if f.get("name") in names or f.get("stored_name") in names or f.get("path") in names]
        parsed = []
        for item in candidates:
            path = Path(item["path"])
            try:
                parsed.append(await self.parse_file(path))
            except Exception as exc:
                logger.exception("Failed to parse %s", path)
                parsed.append({"file": str(path), "status": "failed", "error": f"文件解析失败：{exc}"})
        self.storage.save_json(run_id, "parsed/parsed_documents.json", parsed)
        return parsed

    async def parse_file(self, path: Path) -> dict[str, Any]:
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            return await self.mineru.parse(str(path))
        if suffix in {".md", ".markdown", ".txt"}:
            text = path.read_text(encoding="utf-8", errors="ignore")
            return {"file": str(path), "status": "parsed", "type": "markdown", "title": path.stem, "text": text[:20000], "chunks": self._chunks(text)}
        if suffix == ".csv":
            return self._parse_csv(path)
        if suffix in {".xlsx", ".xls"}:
            return self._parse_excel(path)
        if suffix in {".docx", ".doc"}:
            return self._parse_docx(path)
        return {"file": str(path), "status": "skipped", "reason": "unsupported file type"}

    def _chunks(self, text: str, size: int = 1200) -> list[dict[str, str]]:
        return [{"section": f"chunk-{i//size+1}", "text": text[i:i+size]} for i in range(0, min(len(text), 6000), size)]

    def _parse_csv(self, path: Path) -> dict[str, Any]:
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
            reader = csv.DictReader(f)
            rows = []
            for i, row in enumerate(reader):
                if i >= 10:
                    break
                rows.append(row)
            fields = reader.fieldnames or []
        return {"file": str(path), "status": "parsed", "type": "csv", "title": path.stem, "columns": fields, "sample_rows": rows, "row_sample_count": len(rows)}

    def _parse_excel(self, path: Path) -> dict[str, Any]:
        try:
            from openpyxl import load_workbook
        except Exception:
            return {"file": str(path), "status": "parsed_mock", "type": "excel", "title": path.stem, "note": "openpyxl 未安装，仅保留文件元数据"}
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for ws in wb.worksheets[:5]:
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 8:
                    break
                rows.append([cell for cell in row])
            sheets.append({"sheet": ws.title, "sample_rows": rows})
        return {"file": str(path), "status": "parsed", "type": "excel", "title": path.stem, "sheets": sheets}

    def _parse_docx(self, path: Path) -> dict[str, Any]:
        try:
            import docx
            document = docx.Document(str(path))
            text = "\n".join(p.text for p in document.paragraphs if p.text.strip())
        except Exception:
            try:
                with zipfile.ZipFile(path) as zf:
                    xml = zf.read("word/document.xml")
                root = ElementTree.fromstring(xml)
                text = "\n".join(node.text or "" for node in root.iter() if node.text)
            except Exception as exc:
                return {"file": str(path), "status": "failed", "type": "word", "error": str(exc)}
        return {"file": str(path), "status": "parsed", "type": "word", "title": path.stem, "text": text[:20000], "chunks": self._chunks(text)}
